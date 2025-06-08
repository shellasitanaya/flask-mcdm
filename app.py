from flask import Flask, render_template, request, flash, jsonify, redirect, url_for, send_file
import os
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.comments import Comment
from openpyxl.worksheet.protection import SheetProtection # Correct for SheetProtection
from openpyxl.styles.protection import Protection         # Correct for cell.protection
import json
import io

app = Flask(__name__)
app.secret_key = 'secret-key' 
DOWNLOAD_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'files')

# Daftar kriteria dengan bobot dan tipe
default_criteria = [
    {'kode': 'C1', 'nama': 'IPS', 'tipe': 'Benefit', 'bobot': 0.15, 'description': "Isi dengan IPS yang paling baru (3.01-4.00)"},
    {'kode': 'C2', 'nama': 'Aktif Kemahasiswaan', 'tipe': 'Benefit', 'bobot': 0.10, 'description':'Beri nilai 1-5, dimana:<br>1: Tidak aktif<br>2: Sedikit aktif<br>3: Cukup aktif<br>4: Aktif<br>5: Sangat aktif'},
    {'kode': 'C3', 'nama': 'Kondisi Ekonomi', 'tipe': 'Cost', 'bobot': 0.35, 'description': "Beri nilai 1-5, dimana:<br>1: Tidak berkecukupan<br>2: Sedikit berkecukupan<br>3: Cukup berkecukupan<br>4: Bercukupan<br>5: Sangat bercukupan"},
    {'kode': 'C4', 'nama': 'Semester', 'tipe': 'Benefit', 'bobot': 0.05, 'description': "Isi dengan semester peserta (2-14)"},
    {'kode': 'C5', 'nama': 'Berprestasi', 'tipe': 'Benefit', 'bobot': 0.15, 'description': "Beri nilai 1-5, dimana:<br>1: Tidak berprestasi<br>2: Sedikit berprestasi<br>3: Cukup berprestasi<br>4:Berprestasi<br>5: Sangat berprestasi"},
    {'kode': 'C6', 'nama': 'Motivasi', 'tipe': 'Benefit', 'bobot': 0.20, 'description': "Beri nilai 1-5, dimana:<br>1: Tidak kuat<br>2: Kurang kuat<br>3: Cukup kuat<br>4: Kuat<br>5: Sangat kuat"},
]

# Fungsi untuk melakukan normalisasi matrix berdasarkan tipe kriteria
def normalize(matrix, types):
    # List untuk menyimpan matriks yang sudah dinormalisasi
    normalized = []
    matrix_T = list(zip(*matrix))

    # Iterasi melalui setiap kolom (yang merepresentasikan satu kriteria)
    for j, col in enumerate(matrix_T):
        tipe = types[j]
        col = list(col)
        # Untuk kriteria 'Benefit' (semakin besar semakin baik)
        if tipe == 'Benefit':
            max_val = max(col)
            # Handle case ketikasemua  value nya 0
            norm_col = [x / max_val if max_val != 0 else 0 for x in col]
        # Untuk kriteria 'Cost' (semakin kecil semakin baik)    
        else:  
            min_val = min(col)
            # Handle case ketika min_val nya 0 and x = 0
            norm_col = [min_val / x if x != 0 else 0 for x in col]
        normalized.append(norm_col)

    return list(map(list, zip(*normalized)))

# Fungsi akhir untuk menghitung menggunakan saw
def calculate_saw(matrix, weights, types): 
    # STEP 1 >> Normalisasi Matriks Keputusan
    # Memanggil fungsi normalize() untuk mendapatkan matriks yang sudah dinormalisasi. 
    normalized = normalize(matrix, types)
    weighted_matrix = []
    scores = []

    # Iterasi melalui setiap baris di matriks yang sudah dinormalisasi (setiap baris mewakili satu alternatif)
    for row in normalized:
        # STEP 2 >> Perkalian Matriks Normalisasi dengan Bobot Kriteria
        # Mengalikan setiap nilai yang dinormalisasi dalam baris dengan bobot kriteria yang sesuai.
        # w: bobot kriteria, r: nilai yang sudah dinormalisasi untuk kriteria tersebut.
        weighted_row = [w * r for w, r in zip(weights, row)]
        weighted_matrix.append(weighted_row)

        # STEP 3 >> Penjumlahan Terbobot (Mendapatkan Skor Akhir)
        # Menjumlahkan semua nilai dalam baris yang sudah terbobot untuk mendapatkan skor akhir (Vi) alternatif tersebut.
        scores.append(sum(weighted_row))

    # Mengembalikan tiga hasil:
    # 1. scores: Daftar skor akhir (nilai preferensi) untuk setiap alternatif.
    # 2. normalized: Matriks keputusan yang sudah dinormalisasi.
    # 3. weighted_matrix: Matriks yang sudah dinormalisasi dan dikalikan dengan bobot kriteria.
    return scores, normalized, weighted_matrix

@app.route('/')
def home():
    return render_template('home.html')

# Context processor untuk menyediakan data placeholder dan fungsi get_placeholder_range ke semua template
@app.context_processor
def utility_processor():
    # Fungsi ini mengembalikan string placeholder berdasarkan nama kriteria
    def get_placeholder_range(criteria_name):
        criteria_name_lower = criteria_name.lower()
        if criteria_name_lower == 'ips':
            return '3.01-4'
        elif criteria_name_lower == 'semester':
            return '2-14'
        elif criteria_name_lower in ['aktif kemahasiswaan', 'kondisi ekonomi', 'berprestasi', 'motivasi']:
            return '1-5'
        else:
            return 'Nilai'

    placeholder_data = {
        'ips': '3.01-4',
        'semester': '2-14',
        'aktif kemahasiswaan': '1-5',
        'kondisi ekonomi': '1-5',
        'berprestasi': '1-5',
        'motivasi': '1-5',
    }
    
    return dict(get_placeholder_range=get_placeholder_range, placeholder_data=placeholder_data)
    
# Route untuk halaman SAW 
@app.route('/saw', methods=['GET', 'POST'])
# Inisialisasi variabel untuk menyimpan data dan hasil perhitungan
def saw():
    errors = []
    alternatives = []
    matrix = []
    normalized_matrix = []
    weighted_matrix = []
    scores = []
    ranked = []

    # Jika GET request, gunakan default_criteria.
    # Jika POST request, kriteria akan dibaca dari form_type yang sesuai.
    criteria = default_criteria.copy()

    if request.method == 'POST':
        form_type = request.form.get('form_type')

        # Logika untuk submit form Kriteria 
        if form_type == 'criteria':
            submitted_criteria = []
            criteria_count_str = request.form.get('criteria_count_input')

            if not criteria_count_str:
                flash("Jumlah kriteria tidak terkirim.", "error")
                
            
            try:
                criteria_count = int(criteria_count_str)
            except ValueError:
                flash("Jumlah kriteria tidak valid.", "error")
                criteria_count = 0

            total_bobot = 0
            for i in range(criteria_count):
                name = request.form.get(f'criteria_name_{i}', '').strip()
                bobot_str = request.form.get(f'criteria_weight_{i}', '0')
                tipe = request.form.get(f'criteria_type_{i}', 'Benefit').strip()
                # description = request.form.get(f'criteria_description_{i}', '').strip() 


                if not name:
                    errors.append(f"Nama kriteria ke-{i+1} harus diisi.")
                if not bobot_str:
                    errors.append(f"Bobot untuk kriteria '{name or f'ke-{i+1}'}' harus diisi.")
                if not tipe:
                    errors.append(f"Tipe untuk kriteria '{name or f'ke-{i+1}'}' harus dipilih.")
    
                
                try:
                    bobot = float(bobot_str)
                    if not (0 <= bobot <= 1):
                        errors.append(f"Bobot untuk '{name}' harus antara 0 dan 1.")
                    total_bobot += bobot
                except ValueError:
                    bobot = 0
                    errors.append(f"Bobot untuk '{name}' tidak valid.")

                submitted_criteria.append({'kode': f'CC{i+1}', 'nama': name, 'bobot': bobot, 'tipe': tipe})

            if submitted_criteria and abs(total_bobot - 1.0) > 0.001:
                errors.append(f"Total bobot kriteria ({total_bobot:.2f}) harus sama dengan 1.")

            if errors:
                flash(' '.join(errors), 'error')
                criteria = submitted_criteria # Gunakan kriteria yang disubmit (meskipun ada error)
            else:
                criteria = submitted_criteria # Gunakan kriteria baru yang berhasil
                flash('Kriteria berhasil disimpan!', 'success')


        # Logika untuk submit form Input Data Alternatif & Hitung SAW 
        elif form_type == 'saw':
            # Membangun kembali list 'criteria' dari hidden inputs 
            existing_criteria_count_str = request.form.get('existing_criteria_count', '0')
            try:
                existing_criteria_count = int(existing_criteria_count_str)
            except ValueError:
                errors.append("Jumlah kriteria yang disubmit tidak valid.")
                existing_criteria_count = 0
            
            #  Reset criteria untuk dibangun dari hidden inputs
            criteria = []
            for i in range(existing_criteria_count):
                name = request.form.get(f'existing_criteria_name_{i}', '').strip()
                weight_str = request.form.get(f'existing_criteria_weight_{i}', '0').strip()
                type = request.form.get(f'existing_criteria_type_{i}', 'Benefit').strip()
                
                try:
                    weight = float(weight_str)
                except ValueError:
                    weight = 0 # Default jika ada error
                    errors.append(f"Bobot kriteria '{name}' dari hidden input tidak valid.")
                
                criteria.append({'kode': f'C{i+1}', 'nama': name, 'bobot': weight, 'tipe': type})
            # -----------------------------------------------------------------------

            if not criteria:
                errors.append("Kriteria belum ada. Silakan input kriteria dulu.")
            else:
                alt_count_str = request.form.get('alt_count', '0')
                try:
                    alt_count = int(alt_count_str)
                    if alt_count <= 0:
                        errors.append("Jumlah alternatif harus lebih dari 0.")
                except ValueError:
                    errors.append("Jumlah alternatif tidak valid.")
                    alt_count = 0

                if not errors:
                    for i in range(alt_count):
                        alt_name = request.form.get(f'alt_name_{i}', '').strip()
                        # Kalau nama alternatif kosong, gunakan nama default (A1, A2, dst.)
                        if not alt_name:
                            alt_name = f'A{i+1}'
                        alternatives.append(alt_name)
                        row = []
                        for j in range(len(criteria)):
                            val_str = request.form.get(f'value_{i}_{j}', '0').strip()
                            try:
                                val = float(val_str)
                                if val < 0:
                                    errors.append(f"Nilai untuk '{alt_name}' - '{criteria[j]['nama']}' tidak boleh negatif.")
                                    val = 0
                            except ValueError:
                                val = 0
                                errors.append(f"Nilai untuk '{alt_name}' - '{criteria[j]['nama']}' tidak valid.")
                            row.append(val)
                        matrix.append(row)

                    if not alternatives or not matrix or len(alternatives) != alt_count or len(matrix) != alt_count:
                         errors.append("Data alternatif atau matriks tidak lengkap/kosong.")

                    # Jika tidak ada error dan semua data input lengkap, lakukan perhitungan SAW
                    if not errors and alternatives and matrix and criteria:
                        weights = [c['bobot'] for c in criteria]
                        types = [c['tipe'] for c in criteria]
                        try:
                            # Memanggil fungsi calculate_saw untuk mendapatkan skor, matriks normalisasi, dan matriks terbobot
                            scores, normalized_matrix, weighted_matrix = calculate_saw(matrix, weights, types)
                            # Mengurutkan alternatif berdasarkan skor tertinggi
                            ranked = sorted(zip(alternatives, scores), key=lambda x: x[1], reverse=True)
                            flash('Perhitungan SAW berhasil!', 'success')
                        except Exception as e:
                            errors.append("Error dalam perhitungan SAW: " + str(e))
                    elif not errors:
                         errors.append("Data input tidak lengkap untuk perhitungan SAW.")

            if errors:
                flash(' '.join(errors), 'error')
            
    # Ini akan dieksekusi untuk GET request dan POST request setelah diproses.
    # Variabel `criteria` akan berisi default_criteria (GET),
    # atau hasil submit kriteria (POST criteria),
    # atau kriteria yang dibangun dari hidden inputs (POST saw).
    return render_template('saw.html',
                           criteria=criteria,
                           alternatives=alternatives,
                           matrix=matrix,
                           normalized_matrix=normalized_matrix,
                           weighted_matrix=weighted_matrix,
                           scores=scores,
                           ranked=ranked)


# dematel
@app.route('/dematel', methods=['GET', 'POST'])
def dematel():
    # Jika user melakukan POST request (perhitungan DEMATEL)
    if request.method == 'POST':
        try:
            # Ambil input label dan jumlah tiap kriteria 
            num_criteria = int(request.json['num_criteria'])
            criteria_labels = request.json.get('criteria_labels', [])

            # Bersihkan label (spasi di awal maupun di akhir), gunakan default jika hasilnya kosong
            # Contoh kriteria default: Kriteria 1, Kriteria 2, dst.
            criteria_labels = [label.strip() or f'Kriteria {i+1}' for i, label in enumerate(criteria_labels)]

            # Memastikan criteria_labels memiliki panjang yang benar (jika ada yang terlewat dari frontend)
            if len(criteria_labels) < num_criteria:
                for i in range(len(criteria_labels), num_criteria):
                    criteria_labels.append(f'Kriteria {i+1}')

            if num_criteria <= 1:
                return jsonify({'success': False, 'message': 'Jumlah kriteria harus lebih dari 1.'}), 400

            # Ambil data matriks awal dari input
            matrix_data = []
            for i in range(num_criteria):
                row = []
                for j in range(num_criteria):
                    val = float(request.json[f'matrix_{i}_{j}'])
                    row.append(val)
                matrix_data.append(row)

            # Membuat matriks awal dari data yang diterima
            initial_matrix = np.array(matrix_data)

            # --- Langkah-langkah Perhitungan DEMATEL ---
            # 1. Normalisasi Matriks
            # a. Mencari jumlah maksimum dari setiap baris
            # b. Normalisasi matriks dengan membagi setiap elemen dengan jumlah maksimum
            max_sum = np.sum(initial_matrix, axis=1).max()

            # jika seluruh input berisi 0
            if max_sum == 0:
                return jsonify({'success': False, 'message': 'Seluruh input tidak boleh 0.'}), 400
            
            # Normalisasi matriks dengan membagi setiap elemen dengan jumlah maksimum
            normalized_matrix = initial_matrix / max_sum

            # 2. Matriks Hubungan Total
            # a. Menghitung matriks hubungan total dengan rumus: T = Y . (I-Y)^-1
            identity_matrix = np.identity(num_criteria)
            inv_part = np.linalg.inv(identity_matrix - normalized_matrix)
            total_relation_matrix = normalized_matrix @ inv_part

            # 3. Hitung D (Ri), R (Ci), Prominence, Causal
            D = np.sum(total_relation_matrix, axis=1)
            R = np.sum(total_relation_matrix, axis=0)

            prominence = D + R      # Ri
            causal = D - R          # Ci

            # --- Hitung Type of Identity (identitas) berdasarkan nilai Ri - Ci ---
            identity_types = []
            for val in causal:
                if val > 0.0001:  # Menggunakan sedikit toleransi untuk floating point
                    identity_types.append('Cause')
                else:
                    identity_types.append('Effect')

            # --- Combine D, R, Prominence, Causal, and Identity Type ---
            # Membuat list of dictionaries untuk tabel summary D, R, Prominence, Causal, dan Identity Type
            combined_summary_data = []
            for i in range(num_criteria):
                combined_summary_data.append({
                    'label': criteria_labels[i],
                    'D': D[i].item(),  
                    'R': R[i].item(),
                    'prominence': prominence[i].item(),
                    'causal': causal[i].item(),
                    'type': identity_types[i]
                })

            # --- Hasilkan JSON untuk dikirim ke frontend ---
            results = {
                'initial_matrix': initial_matrix.tolist(),
                'normalized_matrix': normalized_matrix.tolist(),
                'total_relation_matrix': total_relation_matrix.tolist(),
                'criteria_labels': criteria_labels, 
                'combined_summary_data': combined_summary_data 
            }
            return jsonify({'success': True, 'results': results})

        except ValueError:
            return jsonify({'success': False, 'message': 'Pastikan semua input adalah angka yang valid.'}), 400
        except np.linalg.LinAlgError:
            return jsonify({'success': False, 'message': 'Matriks tidak dapat dibalik. Periksa kembali input Anda (mungkin ada dependensi linear atau nilai yang salah).'}), 400
        except Exception as e:
            return jsonify({'success': False, 'message': f'Terjadi kesalahan tak terduga: {str(e)}'}), 500
        
    else: # GET request (akses halaman awal DEMATEL, tanpa input)
        # Inisialisasi kriteria default saat halaman dimuat (untuk Tabitha The Hopeful Scholarship)
        # Jumlah kriteria, label, dan nilai matriks awal default
        num_criteria = 6 
        initial_criteria_labels = ["IPS", "Aktif Kemahasiswaan", "Kondisi Ekonomi", "Semester", "Berprestasi", "Motivasi"] 
        initial_matrix_values = [
            [0, 1, 1, 4, 4, 1], 
            [2, 0, 1, 1, 1, 1], 
            [4, 4, 0, 4, 4, 4],
            [1, 4, 4, 0, 1, 4],
            [2, 3, 2, 3, 0, 1],
            [1, 1, 1, 1, 1, 0]
        ]
        # Pastikan jumlah label sesuai dengan jumlah kriteria, atau akan ditambahkan otomatis di JS
        if len(initial_criteria_labels) < num_criteria:
            for i in range(len(initial_criteria_labels), num_criteria):
                initial_criteria_labels.append(f'Kriteriaa {i+1}')

        # Menampilkan halaman DEMATEL dengan nilai awal
        return render_template('dematel.html',
                            num_criteria=num_criteria,
                            criteria_labels_json=json.dumps(initial_criteria_labels), 
                            initial_matrix_values=json.dumps(initial_matrix_values))


# fungsi untuk membaca file template (xlsx, xls, atau csv dari pengguna)
@app.route('/read-excel', methods=['POST'])
def phpexample():
     
    file = request.files['excel_file']

    if not file:
        return jsonify({"error": "No selected file"}), 400
    
    filename = file.filename

    if filename.endswith(('.xlsx', '.xls')):
        try:
            df = pd.read_excel(file, header=None)
        except Exception as e:
            return jsonify({"error": f"Error reading Excel file: {e}"}), 500
    elif filename.endswith('.csv'):
        try:
            df = pd.read_csv(file)
        except Exception as e:
            return jsonify({"error": f"Error reading CSV file: {e}"}), 500
    else:
        return jsonify({"error": "Unsupported file format. Please upload .xlsx, .xls, or .csv"}), 400
    
    # mempermudah pembacaan data, hapus kolom pertama jika dematel
    if(request.form.get('method_type') is not None and request.form.get('method_type')=="DEMATEL"):
        df.drop(columns=df.columns[0], inplace=True)
        
    data = df.values.tolist()

    for idx1, row in enumerate(data):
        for idx2, val in enumerate(row):
            if not (idx2 == 0 or idx1 == 0):
                try:
                    converted_val = float(val)

                    # check if the converted float is NaN
                    if np.isnan(converted_val):
                        data[idx1][idx2] = 0.0 # ganti pake 0.0 kalok NaN
                    else:
                        data[idx1][idx2] = converted_val 
                except (ValueError, TypeError):
                    # This 'except' block catches:
                    # - Strings that cannot be converted to float (e.g., "hello", "N/A")
                    # - None values (float(None) raises TypeError)
                    data[idx1][idx2] = 0.0

    # kembalikan hasil pembacaan template
    return jsonify(data)


# fungsi membuat template
@app.route('/download_template', methods=['POST'])
def download_template():

    template_name = request.form.get('template_name')

    wb = Workbook()
    ws = wb.active

    if template_name == "DEMATEL":
        criteria_amount = int(request.form.get('criteria_amount'))
        last_col_idx = 1 + criteria_amount

        ws.append([''] + [f'C{i}' for i in range(1, criteria_amount+1)])

        for i in range(1, last_col_idx):
            ws.append([f'C{i}'] + [0]*criteria_amount)


        for i in range(1, 200):
            # make first row bold
            cell = ws.cell(row=1, column=i+1) 
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center')
            if i < last_col_idx:
                cell.comment = Comment("Ganti nama sesuai dengan kriteria anda.", "Author")
            # for diagonal zeros
            cell = ws.cell(row=i+1, column=i+1) 
            cell.font = Font(bold=True, color="FF0000") 
            if i < last_col_idx:
                cell.comment = Comment("Angka merah tidak perlu diubah", "Author")
            # make first column bold
            referenced_cell_name = f'{get_column_letter(i+1)}1'
            formula_string = f'=IF(ISBLANK({referenced_cell_name}),"",{referenced_cell_name})'

            cell = ws.cell(row=i+1, column=1, value=formula_string)
            
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        
        # grid for usable columns
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))
        

        for r_idx in range(1, last_col_idx+1):
            for c_idx in range(1, last_col_idx+1): 
                cell = ws.cell(row=r_idx, column=c_idx) 
                cell.border = thin_border

        # nama sheets
        ws.title = "Input Data DEMATEL"

        # comment kalok boleh nambah kriteria
        ws.cell(row=1, column=last_col_idx+1).comment = Comment("Tambah kriteria di sini.", "Author")

        
    elif template_name == "SAW":
        criterias = json.loads(request.form.get('criterias')) 

        ws.append(['Nama Alternatif'] + criterias)

        # make it bold
        for cell in ws[1]: 
            cell.font = Font(bold=True)
            cell.alignment = Alignment(wrap_text=True, vertical='center')

        # atur lebar
        ws.column_dimensions['A'].width = 25 

        for i, _ in enumerate(criterias):
            col_letter = get_column_letter(i + 2) 
            ws.column_dimensions[col_letter].width = 10

        
        # grid for usable columns
        thin_border = Border(left=Side(style='thin'), 
                            right=Side(style='thin'), 
                            top=Side(style='thin'), 
                            bottom=Side(style='thin'))


        last_col_idx = 1 + len(criterias) 

        for r_idx in range(1, 100 + 1): # r_idx goes from 1 (header) to total_data_rows
            for c_idx in range(1, last_col_idx + 1): # c_idx goes from 1 (col A) to last_col_idx
                cell = ws.cell(row=r_idx, column=c_idx) # Get the specific cell object
                cell.border = thin_border

        # nama sheets
        ws.title = "Input Data SAW"
    else:
        return jsonify({"error": "No template found for this method."}), 400
    
    output = io.BytesIO() 
    wb.save(output)      
    output.seek(0)

    filename = f'template_{template_name}.xlsx'

    # return file template ke frontend
    return send_file(output,
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', 
                    as_attachment=True,
                    download_name=filename)

if __name__ == '__main__':
    app.run(debug=True)